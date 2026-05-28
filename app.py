from flask import Flask, request, render_template, send_file, jsonify
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from PIL import Image as PILImage
import io, os, json, shutil
from datetime import datetime

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, 'template.xlsx')
SAVE_DIR = os.path.join(BASE_DIR, 'saved_data')
META_FILE = os.path.join(SAVE_DIR, 'meta.json')
os.makedirs(SAVE_DIR, exist_ok=True)

LVSBS = [
    {'id': '1', 'name': 'SS5-1-LVSB1', 'sheet': '1_SS5-1-LVSB1'},
    {'id': '2', 'name': 'SS5-1-LVSB2', 'sheet': '2_SS5-1-LVSB2'},
    {'id': '3', 'name': 'SS5-2-LVSB1', 'sheet': '3_SS5-2-LVSB1'},
    {'id': '4', 'name': 'SS5-2-LVSB2', 'sheet': '4_SS5-2-LVSB2'},
    {'id': '5', 'name': 'SS5-3-LVSB1', 'sheet': '5_SS5-3-LVSB1'},
    {'id': '6', 'name': 'SS5-3-LVSB2', 'sheet': '6_SS5-3-LVSB2'},
    {'id': '7', 'name': 'SS5-4-LVSB1', 'sheet': '7_SS5-4-LVSB1'},
    {'id': '8', 'name': 'SS5-4-LVSB2', 'sheet': '8_SS5-4-LVSB2'},
    {'id': '9', 'name': 'SS5-4-LVSB3', 'sheet': '9_SS5-4-LVSB3'},
    {'id': '10', 'name': 'SS5-5-LVSB1', 'sheet': '10_SS5-5-LVSB1'},
    {'id': '11', 'name': 'SS5-5-LVSB2', 'sheet': '11_SS5-5-LVSB2'},
    {'id': '12', 'name': 'SS5-5-LVSB3', 'sheet': '12_SS5-5-LVSB3'},
    {'id': '13', 'name': 'SS6-1-LVSB1', 'sheet': '13_SS6-1-LVSB1'},
    {'id': '14', 'name': 'SS6-1-LVSB2', 'sheet': '14_SS6-1-LVSB2'},
    {'id': '15', 'name': 'SS6-2-LVSB1', 'sheet': '15_SS6-2-LVSB1'},
    {'id': '16', 'name': 'SS6-2-LVSB2', 'sheet': '16_SS6-2-LVSB2'},
    {'id': '17', 'name': 'SS6-3-LVSB1', 'sheet': '17_SS6-3-LVSB1'},
    {'id': '18', 'name': 'SS6-3-LVSB2', 'sheet': '18_SS6-3-LVSB2'},
    {'id': '19', 'name': 'SS6-4-LVSB1', 'sheet': '19_SS6-4-LVSB1'},
    {'id': '20', 'name': 'SS6-4-LVSB2', 'sheet': '20_SS6-4-LVSB2'},
    {'id': '21', 'name': 'SS6-4-LVSB3', 'sheet': '21_SS6-4-LVSB3'},
    {'id': '22', 'name': 'SS6-5-LVSB1', 'sheet': '22_SS6-5-LVSB1'},
    {'id': '23', 'name': 'SS6-5-LVSB2', 'sheet': '23_SS6-5-LVSB2'},
    {'id': '24', 'name': 'SS6-5-LVSB3', 'sheet': '24_SS6-5-LVSB3'},
    {'id': '25', 'name': 'SS7-1-LVSB1', 'sheet': '25_SS7-1-LVSB1'},
    {'id': '26', 'name': 'SS7-1-LVSB2', 'sheet': '26_SS7-1-LVSB2'},
    {'id': '27', 'name': 'SS7-1-LVSB3', 'sheet': '27_SS7-1-LVSB3'},
    {'id': '28', 'name': 'SS7-1-LVSB4', 'sheet': '28_SS7-1-LVSB4'},
    {'id': '29', 'name': 'SS7-2-LVSB1', 'sheet': '29_SS7-2-LVSB1'},
    {'id': '30', 'name': 'SS7-2-LVSB2', 'sheet': '30_SS7-2-LVSB2'},
    {'id': '31', 'name': 'SS7-2-LVSB3', 'sheet': '31_SS7-2-LVSB3'},
    {'id': '32', 'name': 'SS7-2-LVSB4', 'sheet': '32_SS7-2-LVSB4'},
    {'id': '33', 'name': 'SS7-3-LVSB1', 'sheet': '33_SS7-3-LVSB1'},
    {'id': '34', 'name': 'SS7-3-LVSB2', 'sheet': '34_SS7-3-LVSB2'},
    {'id': '35', 'name': 'SS7-3-LVSB3', 'sheet': '35_SS7-3-LVSB3'},
    {'id': '36', 'name': 'SS7-3-LVSB4', 'sheet': '36_SS7-3-LVSB4'},
    {'id': '37', 'name': 'SS7-4-LVSB1', 'sheet': '37_SS7-4-LVSB1'},
    {'id': '38', 'name': 'SS7-4-LVSB2', 'sheet': '38_SS7-4-LVSB2'},
    {'id': '39', 'name': 'SS7-4-LVSB3', 'sheet': '39_SS7-4-LVSB3'},
    {'id': '40', 'name': 'SS7-4-LVSB4', 'sheet': '40_SS7-4-LVSB4'},
    {'id': '41', 'name': 'SU1-1-LVSB1', 'sheet': '41_SU1-1-LVSB1'},
    {'id': '42', 'name': 'SU1-1-LVSB2', 'sheet': '42_SU1-1-LVSB2'},
    {'id': '43', 'name': 'SU1-1-LVSB3', 'sheet': '43_SU1-1-LVSB3'},
    {'id': '44', 'name': 'SU1-1-LVSB4', 'sheet': '44_SU1-1-LVSB4'},
    {'id': '45', 'name': 'SU1-2-LVSB1', 'sheet': '45_SU1-2-LVSB1'},
    {'id': '46', 'name': 'SU1-2-LVSB2', 'sheet': '46_SU1-2-LVSB2'},
    {'id': '47', 'name': 'SU1-2-LVSB3', 'sheet': '47_SU1-2-LVSB3'},
    {'id': '48', 'name': 'SU1-2-LVSB4', 'sheet': '48_SU1-2-LVSB4'},
    {'id': '49', 'name': 'SU1-3-LVSB1', 'sheet': '49_SU1-3-LVSB1'},
    {'id': '50', 'name': 'SU1-3-LVSB2', 'sheet': '50_SU1-3-LVSB2'},
    {'id': '51', 'name': 'SU1-3-LVSB3', 'sheet': '51_SU1-3-LVSB3'},
    {'id': '52', 'name': 'SU1-3-LVSB4', 'sheet': '52_SU1-3-LVSB4'},
    {'id': '53', 'name': 'SU2-1-LVSB1', 'sheet': '53_SU2-1-LVSB1'},
    {'id': '54', 'name': 'SU2-1-LVSB2', 'sheet': '54_SU2-1-LVSB2'},
    {'id': '55', 'name': 'SU2-1-LVSB3', 'sheet': '55_SU2-1-LVSB3'},
    {'id': '56', 'name': 'SU2-1-LVSB4', 'sheet': '56_SU2-1-LVSB4'},
    {'id': '57', 'name': 'SU2-2-LVSB1', 'sheet': '57_SU2-2-LVSB1'},
    {'id': '58', 'name': 'SU2-2-LVSB2', 'sheet': '58_SU2-2-LVSB2'},
    {'id': '59', 'name': 'SU2-2-LVSB3', 'sheet': '59_SU2-2-LVSB3'},
    {'id': '60', 'name': 'SU2-2-LVSB4', 'sheet': '60_SU2-2-LVSB4'},
    {'id': '61', 'name': 'SU2-3-LVSB1', 'sheet': '61_SU2-3-LVSB1'},
    {'id': '62', 'name': 'SU2-3-LVSB2', 'sheet': '62_SU2-3-LVSB2'},
    {'id': '63', 'name': 'SU2-3-LVSB3', 'sheet': '63_SU2-3-LVSB3'},
    {'id': '64', 'name': 'SU2-3-LVSB4', 'sheet': '64_SU2-3-LVSB4'},
    {'id': '65', 'name': 'Unit Sub 1', 'sheet': '65_Unit Sub 1'},
    {'id': '66', 'name': 'Unit Sub 2', 'sheet': '66_Unit Sub 2'},
    {'id': '67', 'name': 'Field Sub 1', 'sheet': '67_Field Sub 1'},
    {'id': '68', 'name': 'Field Sub 2', 'sheet': '68_Field Sub 2'},
    {'id': '69', 'name': 'Field Sub 3', 'sheet': '69_Field Sub 3'},
    {'id': '70', 'name': 'Field Sub 4', 'sheet': '70_Field Sub 4'},
    {'id': '71', 'name': 'Field Sub 5', 'sheet': '71_Field Sub 5'},
    {'id': '72', 'name': 'Field Sub 6', 'sheet': '72_Field Sub 6'},
    {'id': '73', 'name': 'Field Sub 8', 'sheet': '73_Field Sub 8'},
    {'id': '74', 'name': 'Field Sub 9', 'sheet': '74_Field Sub 9'},
    {'id': '75', 'name': 'Field Sub 11', 'sheet': '75_Field Sub 11'},
    {'id': '76', 'name': 'Field Sub 14', 'sheet': '76_Field Sub 14'},
    {'id': '77', 'name': 'Field Sub 15', 'sheet': '77_Field Sub 15'},
    {'id': '78', 'name': 'Field Sub 16', 'sheet': '78_Field Sub 16'},
    {'id': '79', 'name': 'ZONE 2 / SS4-LVSB1', 'sheet': '79_ZONE 2 - SS4-LVSB1'},
    {'id': '80', 'name': 'ZONE 2 / SS4-LVSB2', 'sheet': '80_ZONE 2 - SS4-LVSB2'},
    {'id': '81', 'name': 'ZONE 3 / SS3-LVSB1', 'sheet': '81_ZONE 3 - SS3-LVSB1'},
    {'id': '82', 'name': 'ZONE 3 / SS3-LVSB2', 'sheet': '82_ZONE 3 - SS3-LVSB2'},
    {'id': '83', 'name': 'AMF / MDB1', 'sheet': '83_AMF - MDB1'},
    {'id': '84', 'name': 'AMF / MDB2', 'sheet': '84_AMF - MDB2'},
    {'id': '85', 'name': 'Bus Terminal / MDB', 'sheet': '85_Bus Terminal - MDB'},
    {'id': '86', 'name': 'Water Supply / MDB10001 / MCB1', 'sheet': '86_Water Supply - MDB10001 - MC'},
    {'id': '87', 'name': 'Water Supply / MDB10001 / MCB2', 'sheet': '87_Water Supply - MDB10001 - MC'},
    {'id': '88', 'name': 'Waste Water / MDB 100001', 'sheet': '88_Waste Water - MDB 100001'},
    {'id': '89', 'name': 'Waste Water / MDB 10B00001', 'sheet': '89_Waste Water - MDB 10B00001'},
    {'id': '90', 'name': 'Waste Water / MDB 10B00002', 'sheet': '90_Waste Water - MDB 10B00002'},
    {'id': '91', 'name': 'Solid Waste / MDB', 'sheet': '91_Solid Waste - MDB'},
    {'id': '92', 'name': 'Polder East / MCC1', 'sheet': '92_Polder East - MCC1'},
    {'id': '93', 'name': 'Polder East / MCC2', 'sheet': '93_Polder East - MCC2'},
    {'id': '94', 'name': 'Polder West / MCC1', 'sheet': '94_Polder West - MCC1'},
    {'id': '95', 'name': 'Polder West / MCC2', 'sheet': '95_Polder West - MCC2'},
    {'id': '96', 'name': 'Main Station / MDB', 'sheet': '96_Main Station - MDB'},
    {'id': '97', 'name': 'Fire Traning / MDB1', 'sheet': '97_Fire Traning - MDB1'},
    {'id': '98', 'name': 'AIMS / MDB1', 'sheet': '98_AIMS - MDB1'},
    {'id': '99', 'name': 'AIMS / MDB2', 'sheet': '99_AIMS - MDB2'},
    {'id': '100', 'name': 'AOB / MDB1', 'sheet': '100_AOB - MDB1'},
    {'id': '101', 'name': 'AOB / MDB2', 'sheet': '101_AOB - MDB2'},
    {'id': '102', 'name': 'อาคารพักเวร / MDB', 'sheet': '102_อาคารพักเวร - MDB'},
    {'id': '103', 'name': 'CE / MDB-CE1', 'sheet': '103_CE - MDB-CE1'},
    {'id': '104', 'name': 'CI / MDB-CI1', 'sheet': '104_CI - MDB-CI1'},
    {'id': '105', 'name': 'BC1 / MDB', 'sheet': '105_BC1 - MDB'},
    {'id': '106', 'name': 'BC2-P2 / MDB-P2-A', 'sheet': '106_BC2-P2 - MDB-P2-A'},
    {'id': '107', 'name': 'BC2-P2 / MDB-P2-B', 'sheet': '107_BC2-P2 - MDB-P2-B'},
    {'id': '108', 'name': 'P1 / MDB-P1-A', 'sheet': '108_P1 - MDB-P1-A'},
    {'id': '109', 'name': 'P1 / MDB-P1-B', 'sheet': '109_P1 - MDB-P1-B'},
    {'id': '110', 'name': 'P1 / CL', 'sheet': '110_P1 - CL'},
    {'id': '111', 'name': 'AO-1 / MDB-AO-A', 'sheet': '111_AO-1 - MDB-AO-A'},
    {'id': '112', 'name': 'AO-1 / MDB-AO-B', 'sheet': '112_AO-1 - MDB-AO-B'},
    {'id': '113', 'name': 'AO-2 / MDB-AO-A', 'sheet': '113_AO-2 - MDB-AO-A'},
    {'id': '114', 'name': 'AO-2 / MDB-AO-B', 'sheet': '114_AO-2 - MDB-AO-B'},
    {'id': '115', 'name': 'AO-3 / MDB-AO-A', 'sheet': '115_AO-3 - MDB-AO-A'},
    {'id': '116', 'name': 'AO-3 / MDB-AO-B', 'sheet': '116_AO-3 - MDB-AO-B'},
    {'id': '117', 'name': 'AO-4 / MDB-AO-A', 'sheet': '117_AO-4 - MDB-AO-A'},
    {'id': '118', 'name': 'AO-4 / MDB-AO-B', 'sheet': '118_AO-4 - MDB-AO-B'},
    {'id': '119', 'name': 'S1 / MDB-S1A', 'sheet': '119_S1 - MDB-S1A'},
    {'id': '120', 'name': 'S1 / MDB-S1B', 'sheet': '120_S1 - MDB-S1B'},
    {'id': '121', 'name': 'Ware House1 / MDB-WH-A', 'sheet': '121_Ware House1 - MDB-WH-A'},
    {'id': '122', 'name': 'Ware House1 / MDB-WH-B', 'sheet': '122_Ware House1 - MDB-WH-B'},
    {'id': '123', 'name': 'Ware House2 / MDB-WH-A', 'sheet': '123_Ware House2 - MDB-WH-A'},
    {'id': '124', 'name': 'Ware House2 / MDB-WH-B', 'sheet': '124_Ware House2 - MDB-WH-B'},
    {'id': '125', 'name': 'Ware House3 / MDB-WH-A', 'sheet': '125_Ware House3 - MDB-WH-A'},
    {'id': '126', 'name': 'Ware House3 / MDB-WH-B', 'sheet': '126_Ware House3 - MDB-WH-B'},
    {'id': '127', 'name': 'Ware House4 / MDB-WH-A', 'sheet': '127_Ware House4 - MDB-WH-A'},
    {'id': '128', 'name': 'Ware House4 / MDB-WH-B', 'sheet': '128_Ware House4 - MDB-WH-B'},
    {'id': '129', 'name': 'SS3-5-VVIP-UDBR', 'sheet': '129_SS3-5-VVIP-UDBR'},
]

SLOT_LAYOUT = [
    dict(fr_row=9,  fr_col=0,  fr_rowOff=168965, fr_colOff=341244, to_row=19, to_col=7,  to_rowOff=86590,  to_colOff=278000, w=981, h=981),
    dict(fr_row=9,  fr_col=7,  fr_rowOff=168965, fr_colOff=402949, to_row=19, to_col=14, to_rowOff=86590,  to_colOff=368115, w=981, h=981),
    dict(fr_row=19, fr_col=0,  fr_rowOff=151646, fr_colOff=341244, to_row=29, to_col=7,  to_rowOff=69272,  to_colOff=278000, w=981, h=981),
    dict(fr_row=19, fr_col=7,  fr_rowOff=151646, fr_colOff=402949, to_row=29, to_col=14, to_rowOff=69272,  to_colOff=368115, w=981, h=981),
    dict(fr_row=29, fr_col=0,  fr_rowOff=151646, fr_colOff=341244, to_row=39, to_col=7,  to_rowOff=69272,  to_colOff=278000, w=981, h=981),
    dict(fr_row=29, fr_col=7,  fr_rowOff=151646, fr_colOff=402949, to_row=39, to_col=14, to_rowOff=69272,  to_colOff=368115, w=981, h=981),
    dict(fr_row=41, fr_col=0,  fr_rowOff=173039, fr_colOff=341243, to_row=51, to_col=7,  to_rowOff=90664,  to_colOff=277999, w=981, h=981),
    dict(fr_row=41, fr_col=7,  fr_rowOff=173039, fr_colOff=402948, to_row=51, to_col=14, to_rowOff=90664,  to_colOff=368114, w=981, h=981),
    dict(fr_row=51, fr_col=0,  fr_rowOff=155720, fr_colOff=341243, to_row=61, to_col=7,  to_rowOff=73346,  to_colOff=277999, w=981, h=981),
    dict(fr_row=51, fr_col=7,  fr_rowOff=155720, fr_colOff=402948, to_row=61, to_col=14, to_rowOff=73346,  to_colOff=368114, w=981, h=981),
    dict(fr_row=61, fr_col=0,  fr_rowOff=155720, fr_colOff=341243, to_row=71, to_col=7,  to_rowOff=73346,  to_colOff=277999, w=981, h=981),
    dict(fr_row=61, fr_col=7,  fr_rowOff=155720, fr_colOff=402948, to_row=71, to_col=14, to_rowOff=73346,  to_colOff=368114, w=981, h=981),
]

def load_meta():
    if os.path.exists(META_FILE):
        with open(META_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    # Default to current Thai month/year
    now = datetime.now()
    months_th = ['มกราคม','กุมภาพันธ์','มีนาคม','เมษายน','พฤษภาคม','มิถุนายน',
                 'กรกฎาคม','สิงหาคม','กันยายน','ตุลาคม','พฤศจิกายน','ธันวาคม']
    return {'month_th': months_th[now.month-1], 'year_th': str(now.year+543), 'lvsbs': {}}

def save_meta(meta):
    with open(META_FILE, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

def img_dir(lid):
    d = os.path.join(SAVE_DIR, f'lvsb_{lid}')
    os.makedirs(d, exist_ok=True)
    return d

def resize_image_square(img_bytes, size=981):
    img = PILImage.open(io.BytesIO(img_bytes))
    if img.mode in ('RGBA', 'P'):
        img = img.convert('RGB')
    w, h = img.size
    m = min(w, h)
    img = img.crop(((w-m)//2, (h-m)//2, (w-m)//2+m, (h-m)//2+m))
    img = img.resize((size, size), PILImage.LANCZOS)
    out = io.BytesIO()
    img.save(out, format='JPEG', quality=90)
    return out.getvalue()

def insert_image_exact(ws, img_path, slot):
    xl_img = XLImage(img_path)
    xl_img.width  = slot['w']
    xl_img.height = slot['h']
    anchor = TwoCellAnchor()
    anchor.editAs = 'twoCell'
    anchor._from = AnchorMarker(col=slot['fr_col'], colOff=slot['fr_colOff'], row=slot['fr_row'], rowOff=slot['fr_rowOff'])
    anchor.to    = AnchorMarker(col=slot['to_col'], colOff=slot['to_colOff'], row=slot['to_row'], rowOff=slot['to_rowOff'])
    xl_img.anchor = anchor
    ws.add_image(xl_img)

@app.route('/')
def index():
    meta = load_meta()
    return render_template('index.html', lvsbs=LVSBS, meta=meta)

# แทนที่โค้ดบันทึกรูปเดิมด้วยโค้ดนี้:
for f in request.files.getlist("photos"):
    if not f or not f.filename: continue
    if len(saved_imgs) >= 12: break
    img_bytes = resize_image_square(f.read())
    fname = f"{len(saved_imgs):02d}_{f.filename.replace('/','_')}.jpg"
 
    if GDRIVE_FOLDER_ID:
        # บันทึก temp แล้ว upload ไป Drive
        with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
            tmp.write(img_bytes)
            tmp_path = tmp.name
        file_id = upload_to_drive(tmp_path, fname, GDRIVE_FOLDER_ID)
        os.unlink(tmp_path)
        if file_id:
            saved_imgs.append({"fname": fname, "drive_id": file_id})
    else:
        # Fallback: บันทึก Local
        fpath = os.path.join(img_dir(lid), fname)
        with open(fpath, "wb") as out: out.write(img_bytes)
        saved_imgs.append({"fname": fname, "drive_id": None})


@app.route('/delete_image', methods=['POST'])
@app.route("/image/<lid>/<fname>")
def serve_image(lid, fname):
    entry = load_meta().get("lvsbs",{}).get(lid,{})
    img_info = next((i for i in entry.get("images",[]) if i["fname"]==fname), None)
    if not img_info: return "not found", 404
    if img_info.get("drive_id"):
        data = download_from_drive(img_info["drive_id"])
        if data: return send_file(io.BytesIO(data), mimetype="image/jpeg")
    fpath = os.path.join(img_dir(lid), fname)
    if os.path.exists(fpath): return send_file(fpath, mimetype="image/jpeg")
    return "not found", 404
 
@app.route("/delete_image", methods=["POST"])
def delete_image():
    data  = request.get_json()
    lid   = data.get("lid")
fname = data.get("fname")
    meta  = load_meta()
    entry = meta.get("lvsbs",{}).get(lid,{})
    imgs  = entry.get("images",[])
    img_info = next((i for i in imgs if i["fname"]==fname), None)
    if img_info:
        if img_info.get("drive_id"): delete_from_drive(img_info["drive_id"])
        else:
            fpath = os.path.join(img_dir(lid), fname)
            if os.path.exists(fpath): os.remove(fpath)
        imgs.remove(img_info)
    entry["images"] = imgs
    meta["lvsbs"][lid] = entry
    save_meta(meta)
    return jsonify({"ok": True, "img_count": len(imgs)})


@app.route('/get_state')
def get_state():
    meta = load_meta()
    for lid, entry in meta.get('lvsbs', {}).items():
        entry['images'] = [f for f in entry.get('images', [])
                           if os.path.exists(os.path.join(img_dir(lid), f))]
    return jsonify(meta)

@app.route('/image/<lid>/<fname>')
def serve_image(lid, fname):
    fpath = os.path.join(img_dir(lid), fname)
    if not os.path.exists(fpath):
        return 'not found', 404
    return send_file(fpath, mimetype='image/jpeg')

@app.route('/reset', methods=['POST'])
def reset():
    if os.path.exists(SAVE_DIR):
        shutil.rmtree(SAVE_DIR)
    os.makedirs(SAVE_DIR, exist_ok=True)
    return jsonify({'ok': True})

@app.route('/export')
def export():
    meta     = load_meta()
    month_th = meta.get('month_th', '')
    year_th  = meta.get('year_th', '')
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_main = wb['Main']
    ws_main['A1'] = f'แบบแผนงานสำรวจ  LVSB  ประจำเดือน  {month_th}   {year_th}'

    for lvsb in LVSBS:
        lid   = lvsb['id']
        entry = meta.get('lvsbs', {}).get(lid, {})

        # Update Main sheet — C = วันที่, D = kWh (row = id+3)
        row_idx = int(lid) + 3
        date_val = entry.get('date', '')
        kwh_val  = entry.get('kwh', '')
        if date_val:
            try:
                d = datetime.strptime(date_val, '%Y-%m-%d')
                ws_main.cell(row=row_idx, column=3).value = d.strftime('%d/%m/%Y')
            except:
                ws_main.cell(row=row_idx, column=3).value = date_val
        if kwh_val:
            try:    ws_main.cell(row=row_idx, column=4).value = float(kwh_val)
            except: ws_main.cell(row=row_idx, column=4).value = kwh_val

        # Update detail sheet
        sheet_name = lvsb['sheet']
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if date_val:
            try:
                d = datetime.strptime(date_val, '%Y-%m-%d')
                ws['C7'] = d.strftime('%d/%m/%Y')
            except:
                ws['C7'] = date_val

        # Images
        logo = ws._images[0] if ws._images else None
        ws._images = []
        if logo:
            ws._images.append(logo)
        for i, fname in enumerate(entry.get('images', [])[:12]):
            fpath = os.path.join(img_dir(lid), fname)
            if not os.path.exists(fpath):
                continue
            try:
                insert_image_exact(ws, fpath, SLOT_LAYOUT[i])
            except Exception as e:
                print(f"img error {fname}: {e}")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True,
                     download_name=f'Energy_LVSB_{month_th}_{year_th}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    print("\n" + "="*55)
    print("  ⚡ LVSB Energy Report App  (129 ตู้)")
    print("  เปิดเบราว์เซอร์ไปที่: http://localhost:5050")
    print("="*55 + "\n")
    app.run(debug=False, port=5050)
# เพิ่ม imports เหล่านี้ที่ด้านบนของ app.py (ต่อจาก import เดิม)
import base64, tempfile
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from google.oauth2 import service_account
 
GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID", "")
 
def get_drive_service():
    b64 = os.environ.get("GDRIVE_CREDENTIALS_B64", "")
    if not b64: return None
    creds_json = base64.b64decode(b64).decode("utf-8")
    creds_dict = json.loads(creds_json)
    creds = service_account.Credentials.from_service_account_info(
        creds_dict, scopes=["https://www.googleapis.com/auth/drive"])
    return build("drive", "v3", credentials=creds)
 
def upload_to_drive(local_path, filename, folder_id):
    svc = get_drive_service()
    if not svc: return None
    meta = {"name": filename, "parents": [folder_id]}
    media = MediaFileUpload(local_path, mimetype="image/jpeg")
    f = svc.files().create(body=meta, media_body=media, fields="id").execute()
    return f.get("id")
 
def download_from_drive(file_id):
    svc = get_drive_service()
    if not svc: return None
    req = svc.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done: _, done = dl.next_chunk()
buf.seek(0)
    return buf.read()
 
def delete_from_drive(file_id):
    svc = get_drive_service()
    if svc: svc.files().delete(fileId=file_id).execute()
